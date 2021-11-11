import openpyxl
from Utilities.utils_config_reader import configuration_reader


def get_data(sheet_name):
    workbook = openpyxl.load_workbook(configuration_reader("basic configuration", "data_path"))
    sheet = workbook[configuration_reader("basic configuration", sheet_name)]
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    main_list = []

    for i in range(2, total_rows + 1):
        data_list = []
        for j in range(1, total_cols + 1):
            data = sheet.cell(row=i, column=j).value
            data_list.insert(j, data)
        main_list.insert(i, data_list)
    return main_list


